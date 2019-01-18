# import datetime
# import openpyxl

# book = openpyxl.Workbook()
# sheet1 = book.active
# sheet1.title = "첫번째 시트"
# sheet1.cell(row=1, column=1).value = 'Title'

# sheet2 = book.create_sheet()
# sheet2.title = "두번째 시트"
# sheet2['A1'] = datetime.datetime.now()
# sheet2['A2'] = datetime.date.today()

# sheet2.sheet_properties.tabColor = "1072BA"


# sheet3 = book["두번째 시트"]

# print(book.sheetnames)

# book.save("./data/output.xlsx")

# ----------------
import openpyxl
from pprint import pprint

book = openpyxl.load_workbook("./data/meltop100.xlsx")
sheet = book.worksheets[1]

data = []
for r in sheet.rows:
    data.append([ r[0].value, r[1].value, r[3].value ])

del data[0]    # header 제거

data = sorted(data, key=lambda x: x[2], reverse=True)

pprint(data)