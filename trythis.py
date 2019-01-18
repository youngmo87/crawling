import openpyxl
import csv, codecs
from openpyxl.chart import (
    Reference,
    BarChart
)



book = openpyxl.Workbook()
sheet1 = book.active

with open('./data/meltop100.csv', 'r') as tempFile:
    for r in csv.reader(tempFile):
        tempData=sheet1.append(r)
    
book.save('./data/meltop100.xlsx') 


book = openpyxl.load_workbook("./data/meltop100.xlsx")

sheet = book.active
sheet.delete_rows(1, 1)
sheet.delete_rows(sheet.max_row, 1)


for i in [1, 4, 5]:
    for k in range(1, 101):
        tmpCell = sheet.cell(row=k, column=i)
        tmpCell.value = int(tmpCell.value)
        tmpCell.number_format  
book.save("./data/meltop100try.xlsx")

book1 = openpyxl.load_workbook("./data/meltop100.xlsx")

sheet2 = book1.create_sheet()
sheet2.title = "두번째 시트"

for i in range(1, 101):
    # imgFile = './images/1.png'
    imgFile = "./images/" + str(i) + ".png"
    img = openpyxl.drawing.image.Image(imgFile)
    i = "A" + str(i)
    sheet2.add_image(img, i)

book1.save("./data/meltop100try.xlsx")


# datax = Reference(sheet, min_col=2, 
# 		min_row=1, max_col=2, max_row=10)
# categs = Reference(sheet, min_col=1,
# 				 min_row=1, max_row=10)

# chart = BarChart()
# chart.add_data(data=datax)
# chart.set_categories(categs)

# chart.legend = None  # 범례
# chart.varyColors = True
# chart.title = "차트 타이틀"

# sheet.add_chart(chart, "A8")






# book.save("./data/meltop100try.xlsx")
# imgFile = '.images/'

# data = []
# for r in sheet.rows:
#     data.append([ r[0].value, r[1].value, r[3].value ])

# del data[0]    # header 제거

# data = sorted(data, key=lambda x: x[2], reverse=True)
# pprint(data)




# data = []
# for r in sheet.rows:
#     for c in sheet.columns:
#         data.append([ r[c].value])

# del data[0]    # header 제거

# data = sorted(data, key=lambda x: x[2], reverse=True)
# pprint(data)



# book = openpyxl.Workbook()
# sheet1 = book.active
# sheet1.title = "멜론 TryThis"



# # 저장하기
# book.save("./data/output.xlsx")
